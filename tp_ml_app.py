import streamlit as st
import pandas as pd
import numpy as np
import joblib
from io import BytesIO
from openpyxl import Workbook

# Charger le modèle
model = joblib.load('pipeline_model.pkl')

# Fonction pour créer le fichier Excel
def create_excel(name, data, prediction):
    df = pd.DataFrame(data, index=[0])
    df['prediction'] = prediction
    df['name'] = name

    # Créer un fichier Excel en utilisant openpyxl directement
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Prédiction'
    
    # Ajouter les en-têtes
    for col_num, column_title in enumerate(df.columns, 1):
        sheet.cell(row=1, column=col_num, value=column_title)

    # Ajouter les données
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Sauvegarder le fichier Excel dans BytesIO
    workbook.save(output)
    processed_data = output.getvalue()
    return processed_data

# Page d'accueil
def accueil():
    st.image('stitch.jpg', width=200)  # Remplacez 'stitch.jpg' par le chemin de votre logo
    st.title("Bienvenue sur l'application de prédiction des primes d'assurance maladie")
    st.write("""
        Cette application vous permet de prédire les primes d'assurance maladie 
        en fonction de vos informations personnelles.
        Veuillez entrer votre nom ci-dessous et cliquer sur "Démarrer" pour commencer.
    """)
    name = st.text_input("Entrez votre nom")
    if st.button("Démarrer"):
        st.session_state.name = name
        st.session_state.page = "formulaire"

# Page du formulaire
def formulaire():
    st.title(f"Bienvenue {st.session_state.name}!")
    st.write("Veuillez entrer vos informations ci-dessous pour obtenir une prédiction.")
    
    # Collecte des données utilisateur
    age = st.number_input('Âge du principal bénéficiaire', min_value=0, max_value=100, value=25)
    sex = st.selectbox('Sexe de l\'assuré', ['male', 'female'])
    bmi = st.number_input('Indice de masse corporelle (BMI)', min_value=0.0, max_value=100.0, value=25.0)
    children = st.number_input('Nombre d\'enfants couverts par l\'assurance', min_value=0, max_value=10, value=0)
    smoker = st.selectbox('Statut de fumeur', ['yes', 'no'])
    region = st.selectbox('Zone résidentielle', ['northeast', 'southeast', 'southwest', 'northwest'])

    # Préparation des données pour le modèle
    data = {
        'age': age,
        'sex': sex,
        'bmi': bmi,
        'children': children,
        'smoker': smoker,
        'region': region
    }

    df = pd.DataFrame([data])

    # Prédiction
    if st.button('Prédire'):
        prediction = model.predict(df)[0]
        st.write(f"La prime d'assurance maladie prédite est : {prediction:.2f} $")

        # Télécharger les résultats
        excel_data = create_excel(st.session_state.name, data, prediction)
        st.download_button(
            label="Télécharger les résultats en format Excel",
            data=excel_data,
            file_name=f"prediction_{st.session_state.name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_button"
        )

# Navigation entre les pages
if 'page' not in st.session_state:
    st.session_state.page = 'accueil'

if st.session_state.page == 'accueil':
    accueil()
else:
    formulaire()
